# (c)2021  Henrique Moreira

""" libre -- OpenLibre wrapper, wrapping openpyxl for Excel documents
"""

# pylint: disable=missing-function-docstring

import openpyxl
from xlibre.cell import Textual


class GeneralTabular():
    """ General Tabular abstract class """
    def _simple_column(self, alist:list) -> str:
        def better_str(astr:str):
            return astr.replace("\t", " ").replace(" ", "_").strip()

        res = '+'.join([better_str(item.string) for item in alist])
        return res


class Heading(GeneralTabular):
    """ Sheet heading """
    def __init__(self, sht=None):
        self._sheet = sht
        self._headings = {
            "@lines": [],
            "@columns": [],
            "by-column": {},
            "by-name": {},
            "by-column-letter": {},
        }

    def columns(self) -> list:
        return self._headings["@columns"]

    def by_column(self) -> dict:
        return self._headings["by-column"]

    def get_columns(self, by_what:str) -> str:
        return self._headings[by_what]

    def _my_hash(self, result:dict):
        cols = result["@columns"]
        for col_idx, alist in cols:
            result["by-column"][col_idx] = self._simple_column(alist)
        for col_idx, alist in cols:
            name = result["by-column"][col_idx]
            result["by-name"][name] = col_idx

    def parse(self, start_row:int, end_row:int):
        first_col = -1
        fields, cols = [], {}
        y_row = start_row
        if self._headings["@lines"]:
            return False
        while y_row <= end_row:
            x_col = 1
            while x_col < 10 ** 3:	# max 1000 columns
                text = self._sheet.cell(row=y_row, column=x_col)
                if text.value is None:
                    first_col = x_col
                    break
                field = Textual(text)
                fields.append(field)
                if x_col not in cols:
                    cols[x_col] = [field]
                else:
                    cols[x_col].append(field)
                x_col += 1
            self._headings["@lines"].append((y_row, fields))
            fields = []
            y_row += 1
        for x_col in sorted(cols):
            self._headings["@columns"].append((x_col, cols[x_col]))
        self._my_hash(self._headings)
        return True


class Content():
    """ Sheet content """
    def __init__(self, sht=None, desc="", start_content_row=1):
        self._sheet = sht
        self._description = desc
        self._start_row = int(start_content_row)
        self.rows = []

    def start_row(self) -> int:
        return self._start_row

    def parse(self) -> bool:
        if self._sheet is None:
            return False
        idx = 0
        self._init_parse()
        for row in self._sheet.rows:
            idx += 1
            if idx < self._start_row:
                continue
            alist = [Textual(item) for item in row]
            self.rows.append(alist)
        return True

    def _init_parse(self):
        self.rows = []


class Book():
    """ Workbook class """
    def __init__(self, path=""):
        self._path, self._wbk = path, None
        if path:
            self._load_book()
        self.headings = [Heading()]
        self.contents = [Content(desc=path)]

    def workbook(self):
        return self._wbk

    def sheet_names(self) -> list:
        """ Returns the list of sheet names """
        if not self._wbk:
            return []
        return self._wbk.sheetnames

    def sheet_by_index(self, idx1:int):
        """ Returns the sheet indicated by the index, 'idx1', ranging from 1 to n (sheets).
        """
        if idx1 < 1:
            return None
        if not self._wbk or idx1 > len(self._wbk.sheetnames):
            return None
        name = self._wbk.sheetnames[idx1 - 1]
        sht = self._wbk[name]
        return sht

    def parse_headings(self, start_row:int=1, end_row:int=1) -> bool:
        if not self._wbk:
            return False
        if len(self.headings) > 1:
            self.headings = self.headings[:1]
            self.contents = self.contents[:1]
        for sheetname in self._wbk.sheetnames:
            sht = self._wbk[sheetname]
            head = Heading(sht)
            head.parse(start_row, end_row)
            self.headings.append(head)
            self.contents.append(Content(sht, sheetname, end_row+1))
        return True

    def parse(self, idx1:int=0) -> bool:
        """ Parse sheet with index 'idx1'
        """
        if len(self.headings) <= 1:
            self.parse_headings()
        if idx1 < 0:
            return False
        if idx1 > 0:
            return self.contents[idx1].parse()
        all_ok = True
        if idx1 == 0:
            conts = self.contents[1:]
            for cont in conts:
                if not cont.parse():
                    all_ok = False
        return all_ok

    def _load_book(self) -> str:
        self._wbk = openpyxl.load_workbook(self._path, read_only=True)


if __name__ == "__main__":
    print("Import me!")
